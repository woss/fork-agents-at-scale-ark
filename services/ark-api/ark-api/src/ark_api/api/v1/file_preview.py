from typing import Dict, Any, List, Optional
import base64
import io
from pathlib import Path
from fastapi import APIRouter, HTTPException, Body
from pydantic import BaseModel
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/file-preview", tags=["file-preview"])

class FilePreviewRequest(BaseModel):
    content: str  # Base64 encoded file content
    filename: str
    mimeType: Optional[str] = None

class SpreadsheetData(BaseModel):
    sheets: List[Dict[str, Any]]
    metadata: Dict[str, Any]

def parse_excel_file(content_bytes: bytes, filename: str) -> SpreadsheetData:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
        sheets = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column

            # Extract data
            rows = []
            for row_idx in range(1, min(max_row + 1, 1001)):  # Limit to first 1000 rows for preview
                row_data = []
                for col_idx in range(1, min(max_col + 1, 27)):  # Limit to first 26 columns (A-Z)
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    value = cell.value

                    # Convert to JSON-serializable format
                    if value is None:
                        value = ""
                    elif hasattr(value, 'isoformat'):  # datetime
                        value = value.isoformat()
                    else:
                        value = str(value)

                    row_data.append(value)
                rows.append(row_data)

            # Generate column headers (A, B, C, ...)
            columns = [get_column_letter(i) for i in range(1, min(max_col + 1, 27))]

            sheets.append({
                "name": sheet_name,
                "columns": columns,
                "rows": rows,
                "totalRows": max_row,
                "totalColumns": max_col,
                "previewLimited": max_row > 1000 or max_col > 26
            })

        metadata = {
            "fileType": "excel",
            "filename": filename,
            "sheetCount": len(workbook.sheetnames),
            "hasFormulas": any(
                sheet.cell(row=r, column=c).data_type == 'f'
                for sheet in workbook.worksheets
                for r in range(1, min(sheet.max_row + 1, 10))
                for c in range(1, min(sheet.max_column + 1, 10))
            )
        }

        workbook.close()
        return SpreadsheetData(sheets=sheets, metadata=metadata)

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")

def parse_delimited_file(content_bytes: bytes, filename: str, separator: str = ',', file_type: str = 'csv') -> SpreadsheetData:
    """
    Parse CSV or TSV files with a specified separator.

    Args:
        content_bytes: The file content as bytes
        filename: Name of the file
        separator: Delimiter character (',' for CSV, '\t' for TSV)
        file_type: Type identifier for metadata ('csv' or 'tsv')
    """
    try:
        # Try different encodings
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        df = None

        for encoding in encodings:
            try:
                df = pd.read_csv(io.BytesIO(content_bytes), sep=separator, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue

        if df is None:
            raise ValueError(f"Could not decode {file_type.upper()} with any common encoding")

        # Limit preview size
        preview_df = df.head(1000)

        # Convert to list of lists format
        rows = []

        # Add header row
        headers = list(df.columns)
        rows.append(headers)

        # Add data rows
        for _, row in preview_df.iterrows():
            row_data = []
            for value in row:
                if pd.isna(value):
                    value = ""
                elif hasattr(value, 'isoformat'):  # datetime
                    value = value.isoformat()
                else:
                    value = str(value)
                row_data.append(value)
            rows.append(row_data)

        # Generate column letters
        columns = [get_column_letter(i) for i in range(1, len(headers) + 1)]

        sheets = [{
            "name": "Sheet1",
            "columns": columns,
            "rows": rows,
            "totalRows": len(df) + 1,  # +1 for header
            "totalColumns": len(headers),
            "previewLimited": len(df) > 999
        }]

        metadata = {
            "fileType": file_type,
            "filename": filename,
            "sheetCount": 1,
            "hasFormulas": False,
            "encoding": "auto-detected"
        }

        return SpreadsheetData(sheets=sheets, metadata=metadata)

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse {file_type.upper()} file: {str(e)}")

@router.post("/spreadsheet", response_model=SpreadsheetData)
async def preview_spreadsheet(request: FilePreviewRequest = Body(...)) -> SpreadsheetData:
    try:
        # Decode base64 content
        content_bytes = base64.b64decode(request.content)

        # Determine file type from extension or MIME type
        file_ext = Path(request.filename).suffix.lower()

        if file_ext in ['.xlsx', '.xlsm', '.xls']:
            return parse_excel_file(content_bytes, request.filename)
        elif file_ext == '.csv' or (request.mimeType and 'csv' in request.mimeType):
            return parse_delimited_file(content_bytes, request.filename, separator=',', file_type='csv')
        elif file_ext in ['.tsv', '.tab'] or (request.mimeType and 'tab-separated' in request.mimeType):
            return parse_delimited_file(content_bytes, request.filename, separator='\t', file_type='tsv')
        else:
            # Try to detect format by content
            try:
                # Try Excel first
                return parse_excel_file(content_bytes, request.filename)
            except Exception:
                try:
                    # Try CSV
                    return parse_delimited_file(content_bytes, request.filename, separator=',', file_type='csv')
                except Exception:
                    # Try TSV
                    return parse_delimited_file(content_bytes, request.filename, separator='\t', file_type='tsv')

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Unexpected error: {str(e)}")

@router.get("/health")
async def health():
    return {"status": "ok", "service": "file-preview"}